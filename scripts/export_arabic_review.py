#!/usr/bin/env python3
"""
Build a bilingual review workbook for one framework in the AETHER regulatory library.

    python3 scripts/export_arabic_review.py NCA-ECC [outdir]

Produces <outdir>/aether-arabic-review-<code>.xlsx with one row per control:
the English source beside the Arabic that needs checking, OK/Fix dropdowns, and
correction columns. Feed the completed file back through
scripts/apply_arabic_review.py to write the corrections into the seed JSON and
record who signed off.

Requires openpyxl (preinstalled in the AETHER build container).
"""
import json
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
SEED_DIR = ROOT / "supabase" / "seed" / "regulatory-library"

INK = "FF23304A"
BRASS = "FFC79A45"
PAPER = "FFF7F5F0"
EDIT_FILL = "FFFFF7DA"          # cells the reviewer fills in
HEADER_FILL = "FF23304A"
BORDER = Side(style="thin", color="FFD9D5CC")
BOX = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)

COLUMNS = [
    ("Ref", 12),
    ("Domain", 26),
    ("Subdomain", 24),
    ("Title (EN)", 34),
    ("Title (AR) — review this", 40),
    ("Title verdict", 14),
    ("Title correction (AR)", 40),
    ("Requirement (EN)", 62),
    ("Requirement (AR) — review this", 62),
    ("Requirement verdict", 16),
    ("Requirement correction (AR)", 62),
    ("Reviewer notes", 30),
]
EDIT_COLS = {6, 7, 10, 11, 12}   # 1-indexed: verdicts, corrections, notes


def load(code: str):
    for path in sorted(SEED_DIR.glob("*.json")):
        doc = json.loads(path.read_text(encoding="utf-8"))
        if doc.get("framework", {}).get("code", "").upper() == code.upper():
            return path, doc
    raise SystemExit(f"No seed file found for framework code {code!r} in {SEED_DIR}")


def flatten(doc):
    """Yield control dicts in seed order, with domain/subdomain attached."""
    for domain in doc.get("domains", []):
        subs = domain.get("subdomains") or [
            {"en": None, "controls": domain.get("controls", [])}
        ]
        for sub in subs:
            for control in sub.get("controls", []):
                yield {
                    "domain": domain.get("en"),
                    "subdomain": sub.get("en"),
                    **control,
                }


def build(code: str, outdir: Path) -> Path:
    path, doc = load(code)
    fw = doc["framework"]
    controls = list(flatten(doc))

    wb = Workbook()

    # ---------------------------------------------------------------- read me
    info = wb.active
    info.title = "Read me"
    info.sheet_view.showGridLines = False
    info.column_dimensions["A"].width = 26
    info.column_dimensions["B"].width = 96

    def line(row, label, value, bold=False, size=11):
        info.cell(row=row, column=1, value=label).font = Font(
            name="Arial", size=size, bold=True, color=INK
        )
        c = info.cell(row=row, column=2, value=value)
        c.font = Font(name="Arial", size=size, bold=bold, color=INK)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        return row + 1

    info.cell(row=1, column=1, value="AETHER — Arabic review").font = Font(
        name="Arial", size=16, bold=True, color=INK
    )
    r = 3
    r = line(r, "Framework", f'{fw["short_name"]} — {fw["name_en"]}')
    r = line(r, "Code", fw["code"])
    r = line(r, "Regulator", fw["regulator_en"])
    r = line(r, "Controls to review", len(controls))
    r = line(r, "Generated", date.today().isoformat())
    r += 1

    info.cell(row=r, column=1, value="YOUR NAME").font = Font(
        name="Arial", size=11, bold=True, color=INK
    )
    name_cell = info.cell(row=r, column=2)
    name_cell.fill = PatternFill("solid", fgColor=EDIT_FILL)
    name_cell.border = BOX
    name_cell.font = Font(name="Arial", size=11)
    name_row = r
    r += 1
    info.cell(row=r, column=1, value="REVIEW DATE").font = Font(
        name="Arial", size=11, bold=True, color=INK
    )
    date_cell = info.cell(row=r, column=2, value=date.today().isoformat())
    date_cell.fill = PatternFill("solid", fgColor=EDIT_FILL)
    date_cell.border = BOX
    date_cell.font = Font(name="Arial", size=11)
    date_row = r
    r += 2

    for text in [
        "This name is recorded in the product against every row you approve. "
        "It appears in the control detail as “Verified by …”. Only fill it in if you are "
        "willing to stand behind the Arabic below.",
        "",
        "HOW TO REVIEW",
        f"1. Open the “{fw['short_name']}” tab. One row per control.",
        "2. Read the Arabic against the English beside it. You are checking three things: "
        "the meaning matches, the terminology is what a Saudi regulator and a Saudi CISO "
        "actually use, and the register is formal written Arabic.",
        "3. Set each verdict to OK or Fix. Only the yellow cells are yours to edit.",
        "4. When the verdict is Fix, put the corrected Arabic in the correction column. "
        "Leave the correction blank when the verdict is OK.",
        "5. Rows you leave blank stay unverified in the product and show an "
        "“unverified” badge. Partial reviews are fine — send back what you finished.",
        "",
        "WHAT HAPPENS NEXT",
        "The completed file is read back into the framework source, corrections replace the "
        "generated Arabic, and each reviewed row is stamped with your name and the date. "
        "Rows marked Fix without a correction are left unverified and flagged.",
    ]:
        c = info.cell(row=r, column=2, value=text)
        c.font = Font(
            name="Arial",
            size=11,
            bold=text.isupper() and len(text) < 40,
            color=INK,
        )
        c.alignment = Alignment(wrap_text=True, vertical="top")
        info.row_dimensions[r].height = 30 if len(text) > 90 else 15
        r += 1

    r += 1
    info.cell(row=r, column=1, value="PROGRESS").font = Font(
        name="Arial", size=11, bold=True, color=INK
    )
    r += 1
    sheet_ref = f"'{fw['short_name'][:28]}'"
    last = len(controls) + 1
    for label, formula in [
        ("Titles reviewed", f'=COUNTIF({sheet_ref}!F2:F{last},"OK")+COUNTIF({sheet_ref}!F2:F{last},"Fix")'),
        ("Titles to fix", f'=COUNTIF({sheet_ref}!F2:F{last},"Fix")'),
        ("Requirements reviewed", f'=COUNTIF({sheet_ref}!J2:J{last},"OK")+COUNTIF({sheet_ref}!J2:J{last},"Fix")'),
        ("Requirements to fix", f'=COUNTIF({sheet_ref}!J2:J{last},"Fix")'),
        ("Total rows", len(controls)),
    ]:
        info.cell(row=r, column=1, value=label).font = Font(name="Arial", size=11, color=INK)
        c = info.cell(row=r, column=2, value=formula)
        c.font = Font(name="Arial", size=11, bold=True, color=INK)
        r += 1

    # ------------------------------------------------------------- data sheet
    ws = wb.create_sheet(fw["short_name"][:28])
    ws.sheet_view.showGridLines = False

    for i, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=title)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "D2"

    ar_align = Alignment(wrap_text=True, vertical="top", horizontal="right", readingOrder=2)
    en_align = Alignment(wrap_text=True, vertical="top")

    for row_i, c in enumerate(controls, start=2):
        values = [
            c.get("ref"),
            c.get("domain"),
            c.get("subdomain"),
            c.get("title_en"),
            c.get("title_ar"),
            None,
            None,
            c.get("requirement_en"),
            c.get("requirement_ar"),
            None,
            None,
            None,
        ]
        for col_i, value in enumerate(values, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=value)
            cell.font = Font(name="Arial", size=10, color=INK)
            cell.border = BOX
            cell.alignment = ar_align if col_i in (5, 7, 9, 11) else en_align
            if col_i in EDIT_COLS:
                cell.fill = PatternFill("solid", fgColor=EDIT_FILL)
            elif col_i <= 3:
                cell.fill = PatternFill("solid", fgColor=PAPER)
        est = max(len(str(c.get("requirement_en") or "")), len(str(c.get("requirement_ar") or "")))
        ws.row_dimensions[row_i].height = min(150, max(46, (est // 62) * 13 + 33))

    verdict = DataValidation(
        type="list", formula1='"OK,Fix"', allow_blank=True, showDropDown=False
    )
    verdict.error = "Choose OK or Fix."
    verdict.errorTitle = "Verdict"
    ws.add_data_validation(verdict)
    verdict.add(f"F2:F{len(controls) + 1}")
    verdict.add(f"J2:J{len(controls) + 1}")

    ws.auto_filter.ref = f"A1:L{len(controls) + 1}"

    outdir.mkdir(parents=True, exist_ok=True)
    out = outdir / f"aether-arabic-review-{fw['code'].lower()}.xlsx"
    wb.save(out)
    print(
        f"{fw['code']}: {len(controls)} controls → {out}\n"
        f"  reviewer name cell: 'Read me'!B{name_row}   date cell: 'Read me'!B{date_row}\n"
        f"  source: {path.relative_to(ROOT)}"
    )
    return out


if __name__ == "__main__":
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    target = Path(sys.argv[2]) if len(sys.argv) > 2 else ROOT / "review"
    build(sys.argv[1], target)
