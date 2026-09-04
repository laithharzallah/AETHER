#!/usr/bin/env python3
"""
Read a completed Arabic review workbook back into the framework seed JSON.

    python3 scripts/apply_arabic_review.py review/aether-arabic-review-nca-ecc.xlsx
    python3 scripts/apply_arabic_review.py <file.xlsx> --dry-run

Rules, applied per control row:
  * verdict OK                     -> Arabic kept as-is, row counts as reviewed
  * verdict Fix + correction given -> correction replaces the Arabic, row counts as reviewed
  * verdict Fix, correction blank  -> NOT applied, row left unverified, reported as unresolved
  * both verdicts blank            -> row untouched, stays unverified

A control is marked verified only when BOTH its title and its requirement are
resolved. Verified rows are stamped with the reviewer name and date from the
'Read me' sheet; without a reviewer name nothing is marked verified, because an
unsigned verification is not a verification.

After a successful apply, regenerate the seed migration:
    node scripts/build-regulatory-seed.mjs
"""
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SEED_DIR = ROOT / "supabase" / "seed" / "regulatory-library"


def norm(value):
    return str(value).strip() if value is not None else ""


def read_workbook(path: Path):
    wb = load_workbook(path, data_only=True)
    info = wb["Read me"]

    reviewer = ""
    review_date = ""
    code = ""
    for row in info.iter_rows(min_row=1, max_row=40, max_col=2):
        label = norm(row[0].value).upper()
        value = norm(row[1].value) if len(row) > 1 else ""
        if label == "YOUR NAME":
            reviewer = value
        elif label == "REVIEW DATE":
            review_date = value[:10]
        elif label == "CODE":
            code = value

    data_sheet = next(ws for ws in wb.worksheets if ws.title != "Read me")
    rows = {}
    for r in data_sheet.iter_rows(min_row=2, max_col=12):
        ref = norm(r[0].value)
        if not ref:
            continue
        rows[ref] = {
            "title_verdict": norm(r[5].value).upper(),
            "title_correction": norm(r[6].value),
            "req_verdict": norm(r[9].value).upper(),
            "req_correction": norm(r[10].value),
            "notes": norm(r[11].value),
        }
    return code, reviewer, review_date, rows


def find_seed(code: str):
    for path in sorted(SEED_DIR.glob("*.json")):
        doc = json.loads(path.read_text(encoding="utf-8"))
        if doc.get("framework", {}).get("code", "").upper() == code.upper():
            return path, doc
    raise SystemExit(f"No seed file for framework code {code!r}")


def iter_controls(doc):
    for domain in doc.get("domains", []):
        subs = domain.get("subdomains") or [{"controls": domain.get("controls", [])}]
        for sub in subs:
            for control in sub.get("controls", []):
                yield control


def resolve(verdict, correction, current):
    """-> (new_value, resolved, changed, unresolved)"""
    if verdict == "OK":
        return current, True, False, False
    if verdict == "FIX":
        if correction:
            return correction, True, correction != current, False
        return current, False, False, True
    return current, False, False, False


def main():
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    wb_path = Path(sys.argv[1])
    dry = "--dry-run" in sys.argv
    if not wb_path.exists():
        raise SystemExit(f"Not found: {wb_path}")

    code, reviewer, review_date, rows = read_workbook(wb_path)
    if not code:
        raise SystemExit("Could not read the framework code from the 'Read me' sheet.")
    path, doc = find_seed(code)

    if not reviewer:
        print(
            "WARNING: no reviewer name on the 'Read me' sheet.\n"
            "         Corrections will be applied but NO row will be marked verified.\n"
        )

    stats = {
        "rows_in_workbook": len(rows),
        "controls_in_seed": 0,
        "verified": 0,
        "title_corrections": 0,
        "requirement_corrections": 0,
        "unresolved_fix": [],
        "unmatched_refs": [],
        "untouched": 0,
    }

    seen = set()
    for control in iter_controls(doc):
        stats["controls_in_seed"] += 1
        ref = norm(control.get("ref"))
        review = rows.get(ref)
        if review is None:
            stats["untouched"] += 1
            continue
        seen.add(ref)

        title, t_res, t_chg, t_unres = resolve(
            review["title_verdict"], review["title_correction"], control.get("title_ar", "")
        )
        req, r_res, r_chg, r_unres = resolve(
            review["req_verdict"], review["req_correction"], control.get("requirement_ar", "")
        )

        if t_chg:
            control["title_ar"] = title
            stats["title_corrections"] += 1
        if r_chg:
            control["requirement_ar"] = req
            stats["requirement_corrections"] += 1
        if t_unres or r_unres:
            stats["unresolved_fix"].append(ref)
        if review["notes"]:
            control["review_notes"] = review["notes"]

        if t_res and r_res and reviewer:
            control["verified"] = True
            control["verified_by"] = reviewer
            if review_date:
                control["verified_at"] = review_date
            stats["verified"] += 1
        elif not (t_res or r_res):
            stats["untouched"] += 1

    stats["unmatched_refs"] = sorted(set(rows) - seen)

    print(f"Framework : {code}  ({path.relative_to(ROOT)})")
    print(f"Reviewer  : {reviewer or '(none — nothing will be marked verified)'}")
    print(f"Date      : {review_date or '(none)'}")
    print(f"Controls  : {stats['controls_in_seed']} in seed, {stats['rows_in_workbook']} rows in workbook")
    print(f"Verified  : {stats['verified']}")
    print(f"Corrected : {stats['title_corrections']} titles, {stats['requirement_corrections']} requirements")
    print(f"Untouched : {stats['untouched']}")
    if stats["unresolved_fix"]:
        print(
            f"UNRESOLVED: {len(stats['unresolved_fix'])} row(s) marked Fix with no correction "
            f"— left unverified: {', '.join(stats['unresolved_fix'][:20])}"
            + (" …" if len(stats["unresolved_fix"]) > 20 else "")
        )
    if stats["unmatched_refs"]:
        print(
            f"UNMATCHED : {len(stats['unmatched_refs'])} workbook ref(s) not in the seed: "
            f"{', '.join(stats['unmatched_refs'][:20])}"
        )

    if dry:
        print("\n--dry-run: no files written.")
        return

    path.write_text(json.dumps(doc, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"\nWrote {path.relative_to(ROOT)}")
    print("Next: node scripts/build-regulatory-seed.mjs && supabase db push")


if __name__ == "__main__":
    main()
